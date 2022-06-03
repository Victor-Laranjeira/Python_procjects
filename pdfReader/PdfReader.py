import os.path
import re

import requests
import pdfplumber
import pandas as pd
from pathlib import Path
from collections import namedtuple
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import filedialog
from tkinter.messagebox import *
from tkinter.filedialog import *


def abrir():
    file = askopenfilename(defaultextension=".pdf", filetypes=[("Documento Pdf", ".pdf"), ("Todos os arquivos", "*.*")])
    case = False
    leitura(file, case)


def leitura(file, case):
    try:
        window.title(os.path.basename(file))

        negocios_realizados = namedtuple('negocios_realizados', 'data, c_v, Ativo, Qant, Vlr_Comp, Total')

        with pdfplumber.open(file) as pdf:
            page = pdf.pages[0]
            text = page.extract_text()

            new_negocio_re = re.compile(r'N\s[A-Z]\w*\s(C|V)\s[A-Z]\w*\s[a-]\s([A-Z]+)\s([A-Z.]+)?\s([A-Z0-9]+)?\s(\w+)?\s?\s?['
                r'Da-]\s(\d{1,9999})\s(\d{1,999}\,\d{2})\s(\d{1,999}\.\d{3}\,\d{2})')
            data_negocio_re = re.compile(r'(\d{2}/\d{2}/\d{4})')

            line_items = []
            for line in text.split('\n'):
                line = new_negocio_re.match(line)
                for dat in text.split('\n'):
                    dat = data_negocio_re.match(dat)
                    if line and dat:
                        print(line.group(0))
                        data = dat.group(1)
                        c_v = line.group(1)
                        titulo = line.group(2)
                        quantidade = line.group(6)
                        preco = line.group(7)
                        valor_operacao = line.group(8)
                        line_items.append(negocios_realizados(data, c_v, titulo, quantidade, preco, valor_operacao))
            print(line_items, "\n")
            if case:
                return line_items
            else:
                salvar(line_items)

    except FileNotFoundError:
        window.title("PdfReader")
    except Exception:
        showinfo("Erro de exceção", "O programa encontrou um erro ao tentar abrir o arquivo")


def atualizar():
    file = askopenfilename(defaultextension=".pdf", filetypes=[("Documento pdf", ".pdf"), ("Todos os arquivos", "*.*")])
    try:
        case = True
        texto_novo = leitura(file, case)
        file = askopenfilename(defaultextension=".xlsx", filetypes=[("Documento Excel", ".xlsx"), ("Todos os arquivos", "*.*")])
        texto_existente = pd.read_excel(file, index_col=0, header=1)
        window.title(os.path.basename(file))
        file_path = Path(file)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        df = pd.DataFrame(texto_novo)
        df2 = pd.DataFrame(texto_existente)
        frames = [df2, df]
        result = pd.concat(frames)
        print(df, "\n", df2, "\n")
        print(texto_existente)
        print(result)
        result.to_excel(file_path, sheet_name="bolsa", index=False, index_label=False, startrow=1, startcol=1)

    except FileNotFoundError:
        pass
    # IndexError


def salvar(line_items):
    try:
        file = filedialog.asksaveasfilename(initialfile='Sem Título.xlsx',
                                            defaultextension=".xlsx",
                                            filetypes=[("Todos os arquivos", "*.*"),
                                                       ("Documento Excel", "*.xlsx")])
        file_path = Path(file)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        df = pd.DataFrame(line_items)
        df.to_excel(file_path, sheet_name="bolsa", header=True, index=False, index_label=False, merge_cells=False, startrow=1, startcol=1)

        print(df)
        print(df.info)
        print(df.head())
        print(df.index)

        wb = load_workbook(file_path)
        ws = wb.active # Grab the active worksheet
        print(ws.calculate_dimension())
        # ws.merge_cells("H3:I3")
        # wb.save(file_path)

        if file is None:
            return
    except TypeError:
        pass
    except Exception as e:
        window.title("PdfReader")
        showerror("Erro", "Algo deu errado" + str(e))


def sair():
    window.destroy()


def sobre():
    showinfo("Sobre programa", "Este programa extrai dados específicos de uma nota de corretagem e as plota em uma planilha excel.")


window = Tk()
window.title("PdfReader")

window_width = 700
window_height = 500
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()

x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))

window.geometry("{}x{}+{}+{}".format(window_width, window_height, x, y))

menu_bar = Menu(window)
window.config(menu=menu_bar)

file_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Arquivo", menu=file_menu)
file_menu.add_command(label="Abrir", command=abrir)
file_menu.add_command(label="Salvar", command=salvar)
file_menu.add_command(label="Atualizar", command=atualizar)
file_menu.add_separator()
file_menu.add_command(label="Sair", command=sair)

help_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="Ajuda", menu=help_menu)
help_menu.add_command(label="Sobre", command=sobre)


window.mainloop()
