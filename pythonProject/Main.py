# def teste():
#     print("Teste")
#     loop = int(input(print("Sair do loop: ")))
#     return loop
#
#
# while teste() == 0:
#     teste()

###################################################

# Random method

# import random
#
# x = random.randint(1, 10)
# y = random.random()
# print(y)
#
# myList = ['pedra', 'papel', 'tesoura']
# z = random.choice(myList)
# print(z)
#
# cards = [1, 2, 3, 4, 5, 6, 7, 8, 9, "J", "Q", "K", "A"]
# print(cards)
# random.shuffle(cards)

##################################################

# Exception (Try)

# try:
#     numerator = int(input("Enter a number to divide: "))
#     denominator = int(input("Enter a number to divide by: "))
#     result = numerator / denominator
# except ZeroDivisionError as e:
#     print(e)
#     print("You can´t divide by zero.")
# except ValueError as e:
#     print(e)
#     print("Enter only numbers.")
# except Exception as e:
#     print(e)
#     print("Something went wrong.")
# else:
#     print(result)
# finally:
#     print("This will always execute.")

#####################################################

# File detection
# import os
#
# path = "C:\\Users\\Victor\\OneDrive\\Área de Trabalho"
#
# if os.path.exists(path):
#     print("That location exists.")
#     if os.path.isfile(path):
#         print("that is a file.")
#     elif os.path.isdir(path):
#         print("That is a directory.")
# else:
#     print("That location doesn´t exist.")

# Open, read, write, append, copy, move and delete a file
# Open and read
# try:
#     with open('C:\\Users\\Victor\\OneDrive\\Área de Trabalho\\teste.txt') as file:
#         print(file.read())
# except FileNotFoundError:
#     print("That file was not found.")
# except PermissionError as e:
#     print(e)
#     print("Permission denied.")
#
# write and append
# text = "have a nice day!"
# with open('test.txt', 'w') as file: # 'w' to write, 'a' to append
#     file.write(text)
#
# copy
# import shutil
# shutil.copyfile('test.txt', 'C:\\Users\\Victor\\OneDrive\\Área de Trabalho\\teste.txt')
#
# move
# import os
# source = "path.txt"
# destination = "C:\\Users\\Victor\\OneDrive\\Área de Trabalho\\PathMoved.txt"
#
# try:
#     if os.path.exists(destination):
#         print("There is already a file there.")
#     else:
#         os.replace(source, destination)
#         print(source + " was moved")
# except FileNotFoundError:
#     print(source + "was not found")
#
# DELETE
# import os
# import shutil
# path = "test.txt"
#
# try:
#     os.remove(path) # delete a file
#     os.rmdir(path) # delete an empty directory
#     shutil.rmtree(path) # delete a directry containing files
# except FileNotFoundError:
#     print("That file was not found.")
# except PermissionError:
#     print("You do not have permission to delete that.")
# except OSError:
#     print("You cannot delete that using that function.")
# else:
#     print(path + " was deleted")

###########################################################

# MODULES
# import Messages as msg
#
# msg.hello()
# msg.bye()

###########################################################


# Rock, paper, scissors
# import random
#
# while True:
#     choices = ["rock", "paper", "scissors"]
#
#     computer = random.choice(choices)
#     player = None
#
#     while player not in choices:
#         player = input("rock, papper, or scissors?: ").lower()
#
#     if player == computer:
#         print("computer: ", computer)
#         print("player: ", player)
#         print("Tie.")
#     elif player == "scissors":
#         if computer == "rock":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You lose.")
#         if computer == "paper":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You win.")
#
#     elif player == "paper":
#         if computer == "scissors":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You lose.")
#         if computer == "rock":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You win.")
#
#     elif player == "rock":
#         if computer == "scissors":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You win.")
#         if computer == "paper":
#             print("computer: ", computer)
#             print("player: ", player)
#             print("You lose.")
#
#     play_again = input("Play again? (yes/no): ").lower()
#
#     if play_again != "yes":
#         break
#
# print("Bye.")

#########################################################

# from Car import Car
#
# car_1 = Car("chevy", "Corvette", "2021", "blue")
# car_2 = Car("Ford", "Mustang", "2022", "red")
#
# print(car_1.make)
# print(car_1.model)
# print(car_1.year)
# print(car_1.color)
#
# print(car_2.make)
# print(car_2.model)
# print(car_2.year)
# print(car_2.color)
#
# car_1.stop()
# car_2.drive()

#########################################################

# Walrus operator

# foods = list()
# while True:
#     food = input("What food do you like?: ")
#     if food == "quit":
#         break
#     foods.append(food)
#Does the same thing
# foods = list()
# while food := input("What food do you like?: ") != "quit":
#     foods.append(food)

##########################################################

# Function to variable - How to assign a function to a variable
#
# def hello():
#     print("Hello")
#
# hi = hello
# hello()
# hi()
#
# say = print
# say("That works.")

#########################################################

# High order function = 1. accepts a function as an argument
#                          or
#                       2. return a function
# 1.

# def loud(text):
#     return text.upper()
#
# def quiet(text):
#     return text.lower()
#
# def hello(func):
#     text = func("Hello")
#     print(text)
#
# hello(loud)
# hello(quiet)
# # 2.
# def divisor(x):
#     def dividend(y):
#         return y / x
#     return dividend
#
# divide = divisor(2)
# print(divide(10))

######################################################

# Lambda function
# lambda parameters: expression

# double = lambda x: x * 2
# multiply = lambda x, y: x * y
# add = lambda x, y, z: x + y + z
# full_name = lambda first_name, last_name: first_name + " " + last_name
# age_check = lambda age: True if age >= 18 else False
# print(age_check(18))

######################################################

# Sort iterables

# sort() method = used with lists
# sort() dunction = used with iterables

# students = [("Squidward", "F", "60"),
#            ("Sandy", "A", "33"),
#            ("Patrick", "D", "36"),
#            ("Spongebob", "B", "20"),
#            ("Mr.Krabs", "C", "78")]
#
# age = lambda ages: ages[1]
# students.sort(key=age)
#
# for i in students:
#     print(i)

# students = [("Squidward", "F", "60"),
#            ("Sandy", "A", "33"),
#            ("Patrick", "D", "36"),
#            ("Spongebob", "B", "20"),
#            ("Mr.Krabs", "C", "78")]
#
# age = lambda ages: ages[2]
# sorted_students = sorted(students, key=age, reverse=True)
#
# for i in sorted_students:
#     print(i)

#########################################################

# map function
# map() = aplies a function to each item in an iterable (list, tuple, etc)
# map(function, iterable)

# store = [("shirt", 20.00),
#          ("pants", 25.00),
#          ("jacket", 50.00),
#          ("socks", 10.00)]
#
# to_euros = lambda data: (data[0], data[1]*0.82)
# to_dollars = lambda data: (data[0], data[1]/0.82)
#
# store_euros = list(map(to_euros, store))
#
# for i in store_euros:
#     print(i)

#########################################################

# filter() = creates a collection of elements from an iterable for which a function returns true
# filter(function, iterable)

# friends = [("Rachel", 19),
#            ("Monica", 18),
#            ("Phoebe", 17),
#            ("Joey", 16),
#            ("Chandler", 21),
#            ("Ross", 20)]
#
# age = lambda data: data[1] >= 18
# drinking_buddies = list(filter(age, friends))
#
# for i in drinking_buddies:
#     print(i)

######################################################

# reduce() = apply a function to an iterable and reduce it to a single cumulative value.
#            performs function on first two elements and repeats process until 1 value remains.

# reduce(funtion, iterable)

# import functools
#
# letters = ["h", "e", "l", "l", "o"]
# word = functools.reduce(lambda x, y: x + y, letters)
# print(word)
#
# factorial = [5, 4, 3, 2, 1]
# result = functools.reduce(lambda x, y: x * y, factorial)
# print(result)

#####################################################

# list comprehension = a way to create a new list with less syntax.
#                      can mimic certain lambda functions, easier to read
#                      list = [expression for i in iterable]
#                      list = [expression for i in iterable if conditional]
#                      list = [expression (if/else) for i in iterable]

# squares = []
# for i in range(1, 11):
#     squares.append(i * i)
# print(squares)
#
# squares = [i * i for i in range(1, 11)]
# print(squares)
#
# students = [100, 90, 80, 70, 60, 50, 40, 30, 0]
#
# passed_students = list(filter(lambda x: x >= 60, students))
# passed_students = [i for i in students if i >= 60]
# passed_students = [i if i >= 60 else "FAILED" for i in students]
#
# print(passed_students)

#############################################################

# dictionary comprehension = create dictionaries using expression
#                            can replace for loops and certain lambda functions

# dictionary = {key: expression for (key, value) in iterable}
# dictionary = {key: expression for (key, value) in iterable if conditional}
# dictionary = {key: (if/else) for (key, value) in iterable}
# dictionary = {key: unction(value) for (key, value) in iterable}

# cities_in_F = {'New York': 32, 'Boston': 75, 'Los Angeles': 100, 'Chicago': 50}
# cities_in_C = {key: round((value-32)*(5/9)) for (key,value) in cities_in_F.items()}
# print(cities_in_C)
#
# weather = {'New York': "snowing", 'Boston': "sunny", 'Los Angeles': "sunny", 'Chicago': "cloudy"}
# sunny_weather = {key: value for (key, value) in weather.items() if value == "sunny"}
# print(sunny_weather)
#
# cities = {'New York': 32, 'Boston': 75, 'Los Angeles': 100, 'Chicago': 50}
# desc_cities = {key: ("Warm" if value >= 40 else "COLD") for (key, value) in cities.items()}
# print(desc_cities)
#
# def check_temp(value):
#     if value >= 70:
#         return "HOT"
#     elif 69 >= value >= 40:
#         return "WARM"
#     else:
#         return "COLD"
#
# cities = {'New York': 32, 'Boston': 75, 'Los Angeles': 100, 'Chicago': 50}
# desc_cities = {key: check_temp(value) for (key, value) in cities.items()}
# print(desc_cities)

###################################################################

# zip(*iterables) = aggregate elements from two or more iterables (lis, tuples, sets, etc)
#                   creates a zip object with paired elements stored in tuples for each elemtn

# usernames = ["Dude", "Bro", "Mister"]
# passwords = ["p@ssword", "abc123", "guest"]
# login_date = ["1/1/2021", "1/2/2021", "1/3/2021"]
#
# users = zip(usernames, passwords, login_date)
# for i in users:
#     print(i)

##################################################################

# time module

# import time
#
# print(time.ctime(0)) # convert a time expressed in seconds since epoch to a readable string
# #                      epoch = when you computer think time began (reference point)
# print(time.time()) # return current second since epoch
# print(time.ctime(time.time()))
#
# time_object = time.localtime() # local time
# time_object = time.gmtime() # UTC time
# local_time = time.strftime("%B %d %Y %H:%M:%S", time_object)
# print(local_time)
#
# time_string = "20 april, 2020"
# time_object = time.strptime(time_string, "%d %B, %Y")
# print(time_object)
#
# # (year, month, day, hours, minutes, secs, #day of the week, #day of the year, dst)
# time_tuple = (2020, 4, 20, 4, 0, 0, 0, 0, 0)
# time_string = time.mktime(time_tuple)
# print(time_string)

###################################################################

# thread = a flow of execution. Like a separate order of instructions.
#          However each thread takes a turn running to achieve concurrency.
#          GIL = (global interpreter lock),
#          allows only one thread to hold the ontrol of the python interpreter at any time

# cpu bound = program/task spends most of it's time waiting for internal events(cpu intensive)
#             use multiprocessing

# io bound = program/task spends most of it's time waiting for external events(user input/ web scraping)
#            use multiprocessing

# import threading
# import time
#
# def eat_breakfast():
#     time.sleep(3)
#     print("you ate your breakfast")
#
# def drink_coffee():
#     time.sleep(4)
#     print("you drank coffee")
#
# def study():
#     time.sleep(5)
#     print("you finish studying")
#
# x = threading.Thread(target=eat_breakfast, args=())
# x.start()
#
# y = threading.Thread(target=drink_coffee, args=())
# y.start()
#
# z = threading.Thread(target=study, args=())
# z.start()
#
# x.join()
# y.join()
# z.join()
#
# # eat_breakfast()
# # drink_coffee()
# # study()
#
# print(threading.active_count())
# print(threading.enumerate())
# print(time.perf_counter())

################################################################

# daemon thread = a thread that runs in the background, not important for programs to run
#                 you program will not wait for daemon thread to complete before exiting
#                 non-daemon threads cannot normally be killed, stay alive until task is complete

#                 ex. background tasks, garbage collectin, waiting for input, long running process

# import threading
# import time
#
# def timer():
#     print()
#     print()
#     count = 0
#     while True:
#         time.sleep(1)
#         count += 1
#         print("logged in for ", count, " seconds")
#
# x = threading.Thread(target=timer, daemon=True)
# x.start()
#
# x.setDaemon(True)     # Depreciado
# print(x.isDaemon())
#
# answer = input("Do you wish to exit?")

################################################################

# python multiprocessing
# multiprocessing = running tasks in parallel on different cpu cores, bypasses GIL used for threads
#                   multiprocessing = better for cpu bound tasks (heavy cpu usage)
#                   multithreading = better for io bound tasks (waiting around)

# from multiprocessing import Process, cpu_count
# import time
#
# def counter(num):
#     count = 0
#     while count < num:
#         count += 1
#
# def main():
#
#     print(cpu_count())
#
#     a = Process(target=counter, args=(125000000,))
#     b = Process(target=counter, args=(125000000,))
#     c = Process(target=counter, args=(125000000,))
#     d = Process(target=counter, args=(125000000,))
#     e = Process(target=counter, args=(125000000,))
#     f = Process(target=counter, args=(125000000,))
#     g = Process(target=counter, args=(125000000,))
#     h = Process(target=counter, args=(125000000,))
#
#     a.start()
#     b.start()
#     c.start()
#     d.start()
#     e.start()
#     f.start()
#     g.start()
#     h.start()
#
#     a.join()
#     b.join()
#     c.join()
#     d.join()
#     e.join()
#     f.join()
#     g.join()
#     h.join()
#
#     print("Finished in: ", time.perf_counter(), " seconds")
#
# if __name__ == '__main__':
#     main()

#######################################################################

# GUI windows

# from tkinter import *
#
# window = Tk() #instantiate an instance of a window
# window.geometry("420x420")
# window.title("First GUI program")
#
# icon = PhotoImage(file='logo.png')
# window.iconphoto(True, icon)
# window.config(background="#5cfcff")
#
# window.mainloop()

#######################################################################

# Labels in GUI window
# label = an area widget that holds texts and/or an image within a window

# from tkinter import *
#
# window = Tk()
#
# photo = PhotoImage(file="logo.png")
#
# label = Label(window,
#               text="Hello World",
#               font=('Arial', 40, 'italic'),
#               fg='#00FF00',
#               bg='black',
#               relief=RAISED,
#               bd=10,
#               padx=20,
#               pady=20,
#               image=photo,
#               compound='bottom')
# label.pack()
# #label.place(x=0, y=0)
# window.mainloop()

######################################################################

# button in GUI window

# from tkinter import *
#
# count = 0
#
# def click():
#     global count
#     count += 1
#     print(count)
#
# window = Tk()
#
# photo = PhotoImage(file='logo.png')
#
# button = Button(window,
#                 text="Click me!",
#                 command=click,
#                 font=("Arial", 30),
#                 fg="#00FF00",
#                 bg="black",
#                 activeforeground="#00FF00",
#                 activebackground="black",
#                 state=ACTIVE,
#                 image=photo,
#                 compound="bottom")
#
# button.pack()
#
# window.mainloop()

###########################################################################

# entry widget in GUI window
# entry widget = textbook that accepts a single line of user input

# from tkinter import *
#
# def submit():
#     username = entry.get()
#     print("Hello " +username)
#
# def delete():
#     entry.delete(0, END)
#
# def backspace():
#     entry.delete(len(entry.get())-1, END)
#
# window = Tk()
#
# entry = Entry(window,
#               font=("Arial", 50),
#               fg="#00FF00",
#               bg="black")
#
# # entry.insert(0, "Spongebob")
# # entry.config(show="*")
# # entry.config(state=DISABLED)
#
# entry.pack(side=LEFT)
#
# submit_button = Button(window, text="submit", command=submit)
# submit_button.pack(side=RIGHT)
#
# delete_button = Button(window, text="delete", command=delete)
# delete_button.pack(side=RIGHT)
#
# backspace_button = Button(window, text="backspace", command=backspace)
# backspace_button.pack(side=RIGHT)
#
# window.mainloop()

##########################################################################

# checkbox

# from tkinter import *
#
# def display():
#     if x.get() == 1:
#         print("You agree.")
#     else:
#         print("You don´t agree")
#
# window = Tk()
#
# x = IntVar()
#
# python_photo = PhotoImage(file="logo.png")
#
# check_button = Checkbutton(window,
#                            text="I agree to something.",
#                            variable=x,
#                            onvalue=1,
#                            offvalue=0,
#                            command=display,
#                            font=("Arial", 20),
#                            fg="#00FF00",
#                            bg="black",
#                            activeforeground="#00FF00",
#                            activebackground="black",
#                            padx=25,
#                            pady=15,
#                            image=python_photo,
#                            compound="left")
# check_button.pack()
#
# window.mainloop()

##########################################################################

# radiobutton = similar to checkbox, but you only select one from a group

# from tkinter import *
#
# food = ["pizza", "hamburger", "hotdog"]
#
# def order():
#     if x.get() == 0:
#         print("You ordered pizza.")
#     elif x.get() == 1:
#         print("You ordered hamburger.")
#     else:
#         print("You ordered hotdog.")
#
# window = Tk()
#
# pizzaimage = PhotoImage(file="pizza.png")
# hamburgerimage = PhotoImage(file="hamburger.png")
# hotdogimage = PhotoImage(file="hotdog.png")
# foodimages = [pizzaimage, hamburgerimage, hotdogimage]
#
# x = IntVar()
#
# for i in range(len(food)):
#     radiobutton = Radiobutton(window,
#                               text=food[i],
#                               variable=x,
#                               value=i,
#                               padx=25,
#                               font=("Impact", 25),
#                               image=foodimages[i],
#                               compound=LEFT,
#                               indicatoron=False,
#                               width=375,
#                               command=order)
#     radiobutton.pack(anchor=W)
#
# window.mainloop()

#########################################################################

# Scale

# from tkinter import *
#
# def submit():
#     print("The tamperature is " + str(scale.get()) + " degrees celsius.")
#
# window = Tk()
#
# scale = Scale(window,
#               from_=100,
#               to=0,
#               length=300,
#               orient=VERTICAL,
#               font=("Consolas", 16),
#               tickinterval=10,
#               #showvalue=False,
#               resolution=5,
#               troughcolor="#69EAFF",
#               fg="#FF1C00",
#               bg="#111111")
#
# scale.set(((scale['from']-scale['to'])/2)+scale['to'])
# scale.pack()
#
# button = Button(window, text='submit', command=submit)
# button.pack()
#
# window.mainloop()

####################################################################

# listbox = A listing of selectable text items within it´s own container

# from tkinter import *
#
# def submit():
#     food = []
#     for i in listbox.curselection():
#         food.insert(i, listbox.get(i))
#
#     print("You have ordered: ")
#     for i in food:
#         print(i)
#
# def add():
#     listbox.insert(listbox.size(), entrybox.get())
#     listbox.config(height=listbox.size())
#
# def delete():
#     for i in reversed(listbox.curselection()):
#         listbox.delete(i)
#
#     listbox.config(height=listbox.size())
#
# window = Tk()
#
# listbox = Listbox(window,
#                   bg="#f7ffde",
#                   font=("Constantia", 35),
#                   width=12,
#                   selectmode=MULTIPLE)
# listbox.pack()
#
# listbox.insert(1, "Pizza")
# listbox.insert(2, "Pasta")
# listbox.insert(3, "Garlic Bread")
# listbox.insert(4, "Soup")
# listbox.insert(5, "Sushi")
#
# listbox.config(height=listbox.size())
#
# entrybox = Entry(window)
# entrybox.pack()
#
# submitButton = Button(window, text="Submit", command=submit)
# submitButton.pack()
#
# deleteButton = Button(window, text="Delete", command=delete)
# deleteButton.pack()
#
# addButton = Button(window, text="Add", command=add)
# addButton.pack()
#
# window.mainloop()

########################################################################

# message box

# from tkinter import *
# from tkinter import messagebox
#
# def click():
#     # messagebox.showinfo(title="This is an info message box", message="You are a person.")
#     # messagebox.showwarning(title="WARNING!", message="You have a virus.")
#     # messagebox.showerror(title="ERROR!", message="Something went wrong")
#     #
#     # if messagebox.askokcancel(title="Ask ok cancel", message="Do you want to do the thing?"):
#     #     print("You did the thing.")
#     # else:
#     #     print("You canceled the thing.")
#     #
#     # if messagebox.askretrycancel(title="ask ok cancel", message="Do you want retry the thing?"):
#     #     print("You retried the thing")
#     # else:
#     #     print("You canceled the thing")
#     #
#     # if messagebox.askyesno(title="ask yes or no", message="Do you like cake?"):
#     #     print("I like cake too.")
#     # else:
#     #     print("Why do you not like cake?")
#
#     answer = messagebox.askyesnocancel(title="Yes no cancel", message="Do you like to code?")
#     if answer == TRUE:
#         print("You like to code.")
#     elif answer == FALSE:
#         print("You don´t like coding.")
#     else:
#         print("You have dodge the question.")
#
# window = Tk()
#
# button = Button(window, text="Click me", command=click)
# button.pack()
#
# window.mainloop()

#####################################################################

# color chooser

# from tkinter import *
# from tkinter import colorchooser
#
# def click():
#     window.config(bg=colorchooser.askcolor()[1])
#
# window = Tk()
# window.geometry("420x420")
# button = Button(text="Click me", command=click)
# button.pack()
# window.mainloop()

####################################################################

# text widget = functions like a text area, you can enter  multiple lines of text

# from tkinter import *
#
# def submit():
#     input = text.get("1.0", END)
#     print(input)
#
# window = Tk()
# text = Text(window,
#             bg="light yellow",
#             font=("Ink Free", 25),
#             height=8,
#             width=20,
#             padx=20,
#             pady=20,
#             fg="purple")
# text.pack()
# button = Button(window, text="Submit", command=submit)
# button.pack()
# window.mainloop()

#####################################################################

# file dialog = open and read a text file

# from tkinter import *
# from tkinter import filedialog
#
# def openFile():
#     filepath = filedialog.askopenfilename(initialdir="testfile.txt",
#                                           title="",
#                                           filetypes=(("text files", "*.txt"),
#                                           ("all files", "*.*")))
#     file = open(filepath, 'r')
#     print(file.read())
#     file.close()
#
# window = Tk()
# button = Button(window, text="Open", command=openFile)
# button.pack()
# window.mainloop()

###################################################################

# saving files = create and save a file

# from tkinter import *
# from tkinter import filedialog
#
# def saveFile():
#     file = filedialog.asksaveasfile(initialdir="C:\\Users\\Victor\\Desktop",
#                                     defaultextension='.txt',
#                                     filetypes=[
#                                         ("text file", ".txt"),
#                                         ("HTML file", ".html"),
#                                         ("all files", ".*")
#                                     ])
#     if file is None:
#         return
#     filetext = str(text.get(1.0, END))
#     # filetext = input("Enter some text:")
#     file.write(filetext)
#     file.close()
#
# window = Tk()
# button = Button(window, text="save", command=saveFile)
# button.pack()
# text = Text(window)
# text.pack()
# window.mainloop()

####################################################################

# menu bar

# from tkinter import *
#
# def openFile():
#     print("File has been opened")
# def saveFile():
#     print("File has been saved")
# def cut():
#     print("You cut some text")
# def copy():
#     print("you copied some text")
# def paste():
#     print("You pasted some text")
#
# window = Tk()
# menubar = Menu(window)
# window.config(menu=menubar)
#
# fileMenu = Menu(menubar, tearoff=False, font=("Arial", 15))
# menubar.add_cascade(label="File", menu=fileMenu)
# fileMenu.add_command(label="Open", command=openFile)
# fileMenu.add_command(label="Save", command=saveFile)
# fileMenu.add_separator()
# fileMenu.add_command(label="Exit", command=quit)
#
# editMenu = Menu(menubar, tearoff=False, font=("Arial", 15))
# menubar.add_cascade(label="Edit", menu=editMenu)
# editMenu.add_command(label="Cut", command=cut)
# editMenu.add_command(label="Copy", command=copy)
# editMenu.add_command(label="Paste", command=paste)
# window.mainloop()

#######################################################################

# frame = a rectangular container to group and hold widgets

# from tkinter import *
#
# window = Tk()
# frame = Frame(window, bg="black", bd=5, relief=RAISED)
# frame.place(x=10,y=10)
#
# Button(frame, text="W", font=("Arial", 15), width=3).pack(side=TOP)
# Button(frame, text="A", font=("Arial", 15), width=3).pack(side=LEFT)
# Button(frame, text="S", font=("Arial", 15), width=3).pack(side=LEFT)
# Button(frame, text="D", font=("Arial", 15), width=3).pack(side=LEFT)
# window.mainloop()

#########################################################################

# Creating new window

# from tkinter import *
#
# def create_window():
#     new_window = Tk()       # TopLevel() = new window 'on top' of other windows, linked
#                             # Tk() = new independet window
#     old_window.destroy()    # close out of old windows
#
# old_window = Tk()
#
# Button(old_window, text="Create new window", command=create_window).pack()
#
# old_window.mainloop()

########################################################################

# Creating tabs

# from tkinter import *
# from tkinter import ttk
#
# window = Tk()
#
# notebook = ttk.Notebook(window) # widget that manages a collection of windows/displays
#
# tab1 = Frame(notebook) # new frame for tab 1
# tab2 = Frame(notebook) # new frame for tab 2
#
# notebook.add(tab1, text="Tab 1")
# notebook.add(tab2, text="Tab 2")
# notebook.pack(expand=True, fill="both") # expand = expand to fill any space not otherwise
#                                         # fill = fill space on x and y axis
# Label(tab1, text="Hello", width=50, height=25).pack()
# Label(tab2, text="Goodbye", width=50, height=25).pack()
#
# window.mainloop()

# grid() = geometry manager that organizes widgets ina table-like structure in a parent widget

# from tkinter import *
#
# window = Tk()
#
# titleLabel = Label(window, text="Enter you info", font=("Arial", 16)).grid(row=0, column=0, columnspan=2)
#
# firstNameLabel = Label(window, text="First name", width=20).grid(row=1, column=0)
# firstNameEntry = Entry(window).grid(row=1, column=1)
#
# lastNameLabel = Label(window, text="Last name", width=20).grid(row=2, column=0)
# lastNameEntry = Entry(window).grid(row=2, column=1)
#
# emailLabel = Label(window, text="Email", width=20).grid(row=3, column=0)
# emailEntry = Entry(window).grid(row=3, column=1)
#
# button = Button(window, text="Submit").grid(row=4, column=0, columnspan=2)
#
#
# window.mainloop()

###########################################################################################

# Progress bar

# from tkinter import *
# from tkinter.ttk import *
# import time
#
# def start():
#     GB = 100
#     download = 0
#     speed = 1
#     while download<GB:
#         time.sleep(0.05)
#         bar['value'] += (speed/GB)*100
#         download += speed
#         percent.set(str(int((download/GB)*100))+"%")
#         text.set(str(download)+"/"+str(GB)+" GB completed")
#         window.update_idletasks()
#
# window = Tk()
#
# percent = StringVar()
# text = StringVar()
#
# bar = Progressbar(window, orient=HORIZONTAL, length=300)
# bar.pack(pady=10)
#
# percentLabel = Label(window, textvariable=percent).pack()
# taskLabel = Label(window, textvariable=text).pack
#
# button = Button(window, text="download", command=start).pack()
#
# window.mainloop()

#########################################################################################################

# Canvas = widget that is used to draw graphs, plots, images in a window

# from tkinter import *
#
# window = Tk()
#
# canvas = Canvas(window, height=500, width=500)
# # canvas.create_line(0,0, 500, 500, fill="blue", width=5)
# # canvas.create_line(0,500, 500, 0, fill="red", width=5)
# # canvas.create_rectangle(50, 50, 250, 250, fill="purple")
# # points = [250, 0, 500, 500, 0, 500]
# # canvas.create_polygon(points, fill="yellow", outline="black", width=5)
# # canvas.create_arc(0, 0, 500, 500, style=PIESLICE, start=270, width=5)
# canvas.create_arc(0, 0, 500, 500, fill="red", extent=180, width=10)
# canvas.create_arc(0, 0, 500, 500, fill="white", extent=180, start=180, width=10)
# canvas.create_oval(190, 190, 310, 310, fill="white", width=10)
# canvas.pack()
#
# window.mainloop()

########################################################################################################

# Keybord events

# from tkinter import *
#
# def doSomething(event):
#     # print("You pressed: " + event.keysym)
#     label.config(text=event.keysym)
#
# window = Tk()
#
# window.bind("<Key>", doSomething)
#
# label = Label(window, font=("Arial", 50))
# label.pack()
#
# window.mainloop()

######################################################################################################

# Mouse events

# from tkinter import *
#
# def doSomething(event):
#     print("Mouse coordinates: " + str(event.x)+","+str(event.y))
#
# window = Tk()
#
# # window.bind("<Button-1>", doSomething) #left mouse click
# # window.bind("<Button-2>", doSomething) #scroll wheel click
# # window.bind("<Button-3>", doSomething) #rigth mouse click
# # window.bind("<ButtonRelease>", doSomething)
# # window.bind("<Enter>", doSomething) #Enter the window
# # window.bind("<Leave>", doSomething) #Leave the window
# window.bind("<Motion>", doSomething) #Where the mouse moved
#
# window.mainloop()

#######################################################################################################

# Widget dragging/dropping

# from tkinter import *
#
# def drag_start(event):
#     widget = event.widget
#     widget.startX = event.x
#     widget.startY = event.y
#
# def drag_motion(event):
#     widget = event.widget
#     x = widget.winfo_x() - widget.startX + event.x
#     y = widget.winfo_x() - widget.startY + event.y
#     widget.place(x=x, y=y)
#
# window = Tk()
#
# label = Label(window, bg="red", width=10, height=5)
# label.place(x=0, y=0)
#
# label2 = Label(window, bg="blue", width=10, height=5)
# label2.place(x=100, y=100)
#
# label.bind("<Button-1>", drag_start)
# label.bind("<B1-Motion>", drag_motion)
#
# label2.bind("<Button-1>", drag_start)
# label2.bind("<B1-Motion>", drag_motion)
#
# window.mainloop()

#######################################################################################################

# Move images with keys

# from tkinter import *
#
# def move_up(event):
#     label.place(x=label.winfo_x(), y=label.winfo_y()-10)
# def move_down(event):
#     label.place(x=label.winfo_x(), y=label.winfo_y()+10)
# def move_left(event):
#     label.place(x=label.winfo_x()-10, y=label.winfo_y())
# def move_right(event):
#     label.place(x=label.winfo_x()+10, y=label.winfo_y())
#
# window = Tk()
# window.geometry("500x500")
#
# window.bind("<w>", move_up)
# window.bind("<s>", move_down)
# window.bind("<a>", move_left)
# window.bind("<d>", move_right)
#
# label = Label(window, bg="red", width=2, height=2)
# label.place(x=0, y=0)
#
# window.mainloop()

#########################################################################################################

# Animations

# from tkinter import *
# import time
#
# WIDTH = 500
# HEIGHT = 500
# xVelocity = 3
# yVelocity = 2
#
# window = Tk()
#
# canvas = Canvas(window, width=WIDTH, height=HEIGHT)
# canvas.pack()
#
# background_image = PhotoImage(file='space.png')
# background = canvas.create_image(0, 0, image=background_image, anchor=NW)
#
# photo_image = PhotoImage(file='ufo.png')
# my_image = canvas.create_image(0, 0, image=photo_image, anchor=NW)
#
# image_width = photo_image.width()
# image_height = photo_image.height()
#
# while True:
#     coordinates = canvas.coords(my_image)
#     print(coordinates)
#     if(coordinates[0]>=(WIDTH-image_width) or coordinates[0]<0):
#         xVelocity = -xVelocity
#     if (coordinates[1] >= (HEIGHT - image_height) or coordinates[1] < 0):
#         yVelocity = -yVelocity
#     canvas.move(my_image, xVelocity, yVelocity)
#     window.update()
#     time.sleep(0.01)
#
#
# window.mainloop()

###########################################################################################################

# Multiple animations

# from tkinter import *
# from Ball import *
# import time
#
# window = Tk()
#
# WIDTH = 500
# HEIGHT = 500
#
# canvas = Canvas(window, width=WIDTH, height=HEIGHT)
# canvas.pack()
#
# volley_ball = Ball(canvas, 0, 0, 100, 1, 1, "white")
# basket_ball = Ball(canvas, 0, 0, 50, 4, 3, "blue")
# tennis_ball = Ball(canvas, 0, 0, 20, 7, 5, "yellow")
#
# while True:
#     volley_ball.move()
#     basket_ball.move()
#     tennis_ball.move()
#     window.update()
#     time.sleep(0.01)
#
# window.mainloop()

##########################################################################################################

# Clock program

# from tkinter import *
# from time import *
#
# def update():
#     time_string = strftime("%I:%M:%S %p")
#     time_label.config(text=time_string)
#
#     day_string = strftime("%A")
#     day_label.config(text=day_string)
#
#     date_string = strftime("%B %d, %Y")
#     date_label.config(text=date_string)
#
#     window.after(1000, update)
#
# window = Tk()
#
# time_label = Label(window, font=("Arial", 50), fg="#00FF00", bg="black")
# time_label.pack()
#
# day_label = Label(window, font=("Arial", 25))
# day_label.pack()
#
# date_label = Label(window, font=("Arial", 25))
# date_label.pack()
#
# update()
#
# window.mainloop()

#########################################################################################################

# Send an email

# import smtplib
#
# sender = ""
# receiver = ""
# password = ""
# subject = "E-mail teste usando Python"
# body = "Python e-mail"
#
# # Header
# message = F"""From: Snoop Dogg{sender}
# To: Nicholas Cage{receiver}
# Subject: {subject}\n
# {body}
# """
#
# server = smtplib.SMTP("smtp.gmail.com", 587)
# server.starttls()
#
# try:
#     server.login(sender, password)
#     print("Logged in... ")
#     server.sendmail(sender, receiver, message)
#     print("Email has been sent")
#
# except smtplib.SMTPAuthenticationError:
#     print("Unable to sign in")

###########################################################################################################

# Run with command prompt

# run .py file with cmd

# save file as .py (Python file)
# go to command prompt
# navigate to directory w/ your file: cd C:\Users\Victor\PycharmProjects\pythonProject
# invoke python interpreter + script: python Main.py

# print("Hello World")
#
# name = input("What's your name?: ")
#
# print("Hello " + name)

#########################################################################################################

# ******************************************************************
# Python pip
# ******************************************************************

# Inlcuded for Python versions 3.4+
# open command prompt

#       help:                     pip
#       check:                    pip --version
#       update:                   pip install --upgrade
#       installed packages        pip list
#       check outdated packages   pip list --outadated
#       update a package:         pip install "package name" --upgrade
#       install a package:        pip install "package name"

# ******************************************************************

#######################################################################################################

# Calculator program

# from tkinter import *
#
# def button_press(num):
#     global equation_text
#     equation_text = equation_text + str(num)
#     equation_label.set(equation_text)
#
# def equals():
#     global equation_text
#     try:
#         total = str(eval(equation_text))
#         equation_label.set(total)
#         equation_text = total
#
#     except SyntaxError:
#         equation_label.set("syntax error")
#         equation_text = ""
#
#     except ArithmeticError:
#         equation_label.set("Arithmetic error")
#         equation_text = ""
#
# def clear():
#     global equation_text
#     equation_label.set("")
#     equation_text = ""
#
# window = Tk()
# window.title("Calculator program")
# window.geometry("500x500")
#
# equation_text = ""
# equation_label = StringVar()
#
# label = Label(window, textvariable=equation_label, font=("Arial", 20), bg="white", width=20, height=2)
# label.pack()
#
# frame = Frame(window)
# frame.pack()
#
# button1 = Button(frame, text=1, height=4, width=9, font=35, command=lambda: button_press(1))
# button1.grid(row=0, column=0)
#
# button2 = Button(frame, text=2, height=4, width=9, font=35, command=lambda: button_press(2))
# button2.grid(row=0, column=1)
#
# button3 = Button(frame, text=3, height=4, width=9, font=35, command=lambda: button_press(3))
# button3.grid(row=0, column=2)
#
# button4 = Button(frame, text=4, height=4, width=9, font=35, command=lambda: button_press(4))
# button4.grid(row=1, column=0)
#
# button5 = Button(frame, text=5, height=4, width=9, font=35, command=lambda: button_press(5))
# button5.grid(row=1, column=1)
#
# button6 = Button(frame, text=6, height=4, width=9, font=35, command=lambda: button_press(6))
# button6.grid(row=1, column=2)
#
# button7 = Button(frame, text=7, height=4, width=9, font=35, command=lambda: button_press(7))
# button7.grid(row=2, column=0)
#
# button8 = Button(frame, text=8, height=4, width=9, font=35, command=lambda: button_press(8))
# button8.grid(row=2, column=1)
#
# button9 = Button(frame, text=9, height=4, width=9, font=35, command=lambda: button_press(9))
# button9.grid(row=2, column=2)
#
# button0 = Button(frame, text=0, height=4, width=9, font=35, command=lambda: button_press(0))
# button0.grid(row=3, column=0)
#
# plus = Button(frame, text='+', height=4, width=9, font=35, command=lambda: button_press('+'))
# plus.grid(row=0, column=3)
#
# multiply = Button(frame, text='*', height=4, width=9, font=35, command=lambda: button_press('*'))
# multiply.grid(row=1, column=3)
#
# minus = Button(frame, text='-', height=4, width=9, font=35, command=lambda: button_press('-'))
# minus.grid(row=2, column=3)
#
# divide = Button(frame, text='/', height=4, width=9, font=35, command=lambda: button_press('/'))
# divide.grid(row=3, column=3)
#
# equal = Button(frame, text='equal', height=4, width=9, font=35, command=equals)
# equal.grid(row=3, column=2)
#
# decimal = Button(frame, text='.', height=4, width=9, font=35, command=lambda: button_press('.'))
# decimal.grid(row=3, column=1)
#
# clear = Button(window, text='clear', height=4, width=12, font=35, command=clear)
# clear.pack()
#
#
# window.mainloop()

##########################################################################################################

# Text editor program

# import os
# from tkinter import *
# from tkinter import filedialog, colorchooser, font
# from tkinter.messagebox import *
# from tkinter.filedialog import *
#
# def chage_color():
#     color = colorchooser.askcolor(title="Pick a color")
#     text_area.config(fg=color[1])
#
# def change_font(*args):
#     text_area.config(font=(font_name.get(), size_box.get()))
#
# def new_file():
#     window.title("Untitled")
#     text_area.delete(1.0, END)
#
# def open_file():
#     file = askopenfilename(defaultextension=".txt", filetypes=[("All files", "*.*"), ("Text documents", ".txt")])
#
#     try:
#         window.title(os.path.basename(file))
#         text_area.delete(1.0, END)
#         file = open(file, "r")
#         text_area.insert(1.0, file.read())
#
#     except Exception:
#         print("Couldn´t read the file")
#
#     finally:
#         file.close()
#
# def save_file():
#     file = filedialog.asksaveasfilename(initialfile='untitled.txt',
#                                         defaultextension=".txt",
#                                         filetypes=[("All files", "*.*"),
#                                                    ("Text documents", "*.txt")])
#
#     if file is None:
#         return
#     else:
#         try:
#             window.title(os.path.basename(file))
#             file = open(file, "w")
#             file.write(text_area.get(1.0, END))
#
#         except Exception:
#             print("Couldn´t save file")
#
#         finally:
#             file.close()
#
# def cut():
#     text_area.event_generate("<<Cut>>")
#
# def copy():
#     text_area.event_generate("<<Copy>>")
#
# def paste():
#     text_area.event_generate("<<Paste>>")
#
# def about():
#     showinfo("About program", "This program enable you to write whatever you want.")
#
# def quit():
#     window.destroy()
#
# window = Tk()
# window.title("Text editor program")
# file = None
#
# window_width = 500
# window_height = 500
# screen_width = window.winfo_screenwidth()
# screen_height = window.winfo_screenheight()
#
# x = int((screen_width / 2) - (window_width / 2))
# y = int((screen_height / 2) - (window_height / 2))
#
# window.geometry("{}x{}+{}+{}".format(window_width, window_height, x, y))
#
# font_name = StringVar(window)
# font_name.set("Arial")
#
# font_size = StringVar(window)
# font_size.set("25")
#
# text_area = Text(window, font=(font_name.get(), font_size.get()))
#
# scrool_bar = Scrollbar(text_area)
# window.grid_rowconfigure(0, weight=1)
# window.grid_columnconfigure(0, weight=1)
# text_area.grid(sticky=N + E + S + W)
# scrool_bar.pack(side=RIGHT, fill=Y)
# text_area.config(yscrollcommand=scrool_bar.set)
#
# frame = Frame(window)
# frame.grid()
#
# color_button = Button(frame, text="color", command=chage_color)
# color_button.grid(row=0, column=0)
#
# font_box = OptionMenu(frame, font_name, *font.families(), command=change_font)
# font_box.grid(row=0, column=1)
#
# size_box = Spinbox(frame, from_=1, to=100, textvariable=font_size, command=change_font)
# size_box.grid(row=0, column=2)
#
# menu_bar = Menu(window)
# window.config(menu=menu_bar)
#
# file_menu = Menu(menu_bar, tearoff=0)
# menu_bar.add_cascade(label="File", menu=file_menu)
# file_menu.add_command(label="New", command=new_file)
# file_menu.add_command(label="Open", command=open_file)
# file_menu.add_command(label="Save", command=save_file)
# file_menu.add_separator()
# file_menu.add_command(label="Exit", command=quit)
#
# edit_menu = Menu(menu_bar, tearoff=0)
# menu_bar.add_cascade(label="Edit", menu=edit_menu)
# edit_menu.add_command(label="cut", command=cut)
# edit_menu.add_command(label="copy", command=copy)
# edit_menu.add_command(label="paste", command=paste)
#
# help_menu = Menu(menu_bar, tearoff=0)
# menu_bar.add_cascade(label="Help", menu=help_menu)
# help_menu.add_command(label="About", command=about)
#
# window.mainloop()

############################################################################################################

# Tic Tac Toe game

# from tkinter import *
# import random
#
# def next_turn(row, column):
#     global player
#     if buttons[row][column]['text'] == "" and check_winner() is False:
#         if player == players[0]:
#             buttons[row][column]['text'] = player
#             if check_winner() is False:
#                 player = players[1]
#                 label.config(text=(players[1]+" turn"))
#
#             elif check_winner() is True:
#                 label.config(text=(players[0]+" wins"))
#
#             elif check_winner() == "Tie":
#                 label.config(text="Tie!")
#
#         else:
#             buttons[row][column]['text'] = player
#             if check_winner() is False:
#                 player = players[0]
#                 label.config(text=(players[0] + " turn"))
#
#             elif check_winner() is True:
#                 label.config(text=(players[1] + " wins"))
#
#             elif check_winner() == "Tie":
#                 label.config(text="Tie!")
#
#
# def check_winner():
#     for row in range(3):
#         if buttons[row][0]['text'] == buttons[row][1]['text'] == buttons[row][2]['text'] != "":
#             buttons[row][0].config(bg="green")
#             buttons[row][1].config(bg="green")
#             buttons[row][2].config(bg="green")
#             return True
#     for column in range(3):
#         if buttons[0][column]['text'] == buttons[1][column]['text'] == buttons[2][column]['text'] != "":
#             buttons[0][column].config(bg="green")
#             buttons[1][column].config(bg="green")
#             buttons[2][column].config(bg="green")
#             return True
#     if buttons[0][0]['text'] == buttons[1][1]['text'] == buttons[2][2]['text'] != "":
#         buttons[0][0].config(bg="green")
#         buttons[1][1].config(bg="green")
#         buttons[2][2].config(bg="green")
#         return True
#     elif buttons[0][2]['text'] == buttons[1 ][1]['text'] == buttons[2][0]['text'] != "":
#         buttons[0][2].config(bg="green")
#         buttons[1][1].config(bg="green")
#         buttons[2][0].config(bg="green")
#         return True
#     elif empty_spaces() is False:
#         for row in range(3):
#             for column in range(3):
#                 buttons[row][column].config(bg="yellow")
#         return "Tie"
#     else:
#         return False
#
#
# def empty_spaces():
#     spaces = 9
#     for row in range(3):
#         for column in range(3):
#             if buttons[row][column]['text'] != "":
#                 spaces -=1
#     if spaces == 0:
#         return False
#     else:
#         return True
#
# def new_game():
#     global player
#     player = random.choice(players)
#     label.config(text=player+" turn")
#     for row in range(3):
#         for column in range(3):
#             buttons[row][column].config(text="", bg="#F0F0F0")
#
# window = Tk()
# window.title("Tic-Tac-Toe")
# players = ["x", "O"]
# player = random.choice(players)
# buttons = [[0, 0, 0],
#            [0, 0, 0],
#            [0, 0, 0]]
#
# label = Label(text=player + " turn", font=("Arial", 30))
# label.pack(side="top")
#
# reset_button = Button(text="Restart", font=("Arial", 20), command=new_game)
# reset_button.pack(side="top")
#
# frame = Frame(window)
# frame.pack()
#
# for row in range(3):
#     for column in range(3):
#         buttons[row][column] = Button(frame, text="", font=("Arial", 30),
#                                       width=5, height=2,
#                                       command=lambda row=row, column=column: next_turn(row, column))
#         buttons[row][column].grid(row=row, column=column)
#
#
# window.mainloop()

##################################################################################################################

from openpyxl import Workbook, load_workbook
import pandas as pd

wb = load_workbook("Teste.xlsx")
ws = wb.active

# ws1 = wb.create_sheet("Planilha teste") # Criando planilha
# ws.title = "Título teste" # Adicionando título a planilha
# ws.sheet_properties.tabColor = "1072BA" # Cor de fundo da tabela da planilha
print(wb.worksheets) # Vizualizar o nome de todas as tabelas

# for sheet in wb:
#     print(sheet.title)  # Iterando entre planilhas

#source = wb.active
#target = wb.copy_worksheet(source)  # copiando planilha

# c = ws['A4'] # Acessando célula diretamente
# ws['A4'] = 4 # Alterando valores diretamente

#d = ws.cell(row=4, column=1, value=10) # Acessando célula usando notação de linhas e colunas (Row=3, column=5)
#cell_range = ws['B7':'F10'] # Acessando o escopo de células selecionadas

#print(cell_range) # printando as células selecionadas

# for row in ws.iter_rows(min_row=8, max_col=4, max_row=10): # Usando o método Worksheet.iter.rows() para retornar as linhas
#     for cell in row:
#         print(cell)

# for col in ws.iter_cols(min_row=8, max_col=4, max_row=10): # Usando o método Worksheet.iter.rows() para retornar as colunas
#     for cell in col:
#         print(cell)

# for row in ws.values: # Worksheet.values propriedades para acessar apenas os valores
#     if row is None:
#         pass
#     for value in row:
#         print(value)
#
# for row in ws.iter_rows(min_row=2, max_col=9, max_row=8, values_only=True): # Worksheet.iter_rows irá retornar apenas valores
#     print(row)


# if ws.calculate_dimension() is not None:
#     print(ws.calculate_dimension())
#     posicao = ws.calculate_dimension()
#     posicao = list(posicao)
#     if len(posicao) == 4:
#         linha = int(posicao[4])
#     elif len(posicao) == 5:
#         linha = int(posicao[5])
#     else:
#         linha = int(posicao[6])
#     coluna = posicao[0]
#     print(type(linha))
#     linha += 1
#     print(coluna, linha)

df = pd.read_excel('Teste.xlsx', index_col=0, header=0, usecols=[1, 2, 6, 7, 8])
print(df)

# for line in list(posicao):
#     print(posicao[1])


# for col in ws.iter_cols(min_row=2, max_col=9, max_row=8):
#     for cell in col:
#         print(cell)
#
# lista = []
# count = 0
# for line in ws.rows:
#     # count += 1
#     lista = [line]
#     print(line)
#     # print(lista[])



#ws.move_range("A1:I8", rows=3, cols=3)
# ws.merge_cells("L4:M4")
#
#
# wb.save("Teste.xlsx") # Salvar